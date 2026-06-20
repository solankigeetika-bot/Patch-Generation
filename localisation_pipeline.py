#!/usr/bin/env python3
"""
End-to-end localization pipeline for PocketFM shows.

Replaces the complete human verifier workflow:
  Stage 1  — Master ingestion: understand show universe & character bible
  Stage 2  — Proposition review: evaluate each localized name for fit
  Stage 3  — Consistency cascade: ensure canonical decisions apply to all mentions
  Stage 4  — Uniformly flat rule: same source name → always same target name
  Stage 5  — Cross-batch consistency: new episodes must respect prior decisions
             (e.g. if Eps 1-100 use "Shirley"→"Agathe", Eps 101-200 must not use
             "Shirley"→"Charlène")

Usage:
    python localisation_pipeline.py \\
        --input  TOLR_1-100_source.xlsx \\
        --output TOLR_1-100_verified.xlsx \\
        --master TOLR_Master.xlsx \\
        --previous TOLR_prev1.xlsx TOLR_prev2.xlsx \\
        --source-lang de --target-lang fr

    # Rule checks only (no LLM, no network needed):
    python localisation_pipeline.py --input ... --output ... --dry-run

Configuration (never hardcode the key — use environment variables):
    ARGUS_API_KEY     required for LLM proposition review
    ARGUS_BASE_URL    e.g. https://argus.pocketfm.org/api
    ARGUS_MODEL       model id served by Argus (default: claude-opus-4.8)

NOTE: Argus (argus.pocketfm.org) is publicly reachable, so this runs anywhere.
Use --dry-run to skip the LLM step entirely.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ─── Cell fill colours ───────────────────────────────────────────────────────
FILL_CRITICAL = PatternFill("solid", fgColor="FF4444")
FILL_HIGH     = PatternFill("solid", fgColor="FFA500")
FILL_MEDIUM   = PatternFill("solid", fgColor="FFFF00")
FILL_LOW      = PatternFill("solid", fgColor="ADD8E6")
FILL_PASS     = PatternFill("solid", fgColor="90EE90")

SEV_FILL = {
    "critical": FILL_CRITICAL,
    "high":     FILL_HIGH,
    "medium":   FILL_MEDIUM,
    "low":      FILL_LOW,
}


# ─── Language metadata ────────────────────────────────────────────────────────
LANGS = {
    "de": "German",  "fr": "French",    "en": "English",
    "it": "Italian", "tr": "Turkish",   "es": "Spanish",
    "pt": "Portuguese",
}

def lang_name(code: str) -> str:
    return LANGS.get(code, LANGS.get(code.split("-")[0], code))


# ─── Shared helpers ───────────────────────────────────────────────────────────
EMPTY_VALS = {"", "none", "null", "n/a", "-", "undefined", "na", "#n/a", "tbd", "tbc"}
CHAPTER_RE = re.compile(r"^\d+(\s*,\s*\d+)*$")
LOWER_START = re.compile(r"^[a-zà-öø-ÿ]")

KINSHIP = {
    "mama", "maman", "papa", "schwester", "schwägerin", "bruder", "tante", "onkel",
    "belle-sœur", "sœur", "frère", "grand-père", "grand-mère", "chère",
    "großvater", "großmutter", "oma", "opa", "schatz", "liebling",
    "belle-mère", "beau-père", "neveu", "nièce", "cousin", "cousine",
}

VALID_CULTURAL = {
    "LOCALIZED", "KEPT", "ADAPTED", "TRANSLATED", "REMOVED",
    "ADDED", "MODIFIED", "ORIGINAL", "MIXED", "UNIVERSAL",
}

SOURCE_LEAK = {
    "de": re.compile(
        r"\b(der|die|das|und|von|für|ist|bei|mit|auf|des|dem|den|ein|eine|einer|"
        r"einen|einem|eines|nicht|schwester|schwägerin|schatz|bruder|vater|mutter|"
        r"großvater|großmutter|herr|frau|kaiser|konzern|gesellschaft|firma)\b",
        re.IGNORECASE,
    ),
}

# Human-readable tag map (internal kind → LS column tag)
KIND_TAG = {
    "missing_localised":           "MISSING_LOCALISATION",
    "missing_cultural_status":     "MISSING_LOCALISATION",
    "invalid_cultural_status":     "MISSING_LOCALISATION",
    "missing_localisation_reason": "MISSING_LOCALISATION",
    "name_part_mismatch":          "MENTION_MAP_INCONSISTENCY",
    "source_lang_leak":            "ENTITY_INCONSISTENCY",
    "source_name_not_localised":   "ENTITY_INCONSISTENCY",
    "status_mismatch":             "ENTITY_INCONSISTENCY",
    "casing":                      "ENTITY_INCONSISTENCY",
    "master_mismatch":             "CHARACTER_CONTEXT_MISMATCH",
    "cross_character_inconsistency": "CROSS_CHARACTER_INCONSISTENCY",
    "cross_entity_inconsistency":  "CROSS_ENTITY_INCONSISTENCY",
    "mention_map_inconsistency":   "MENTION_MAP_INCONSISTENCY",
    "cross_batch_inconsistency":   "CROSS_CHARACTER_INCONSISTENCY",
    "llm_proposition":             "CHARACTER_CONTEXT_MISMATCH",
    "llm_error":                   "ENTITY_INCONSISTENCY",
}


def is_empty(v) -> bool:
    return str(v or "").strip().lower() in EMPTY_VALS


def norm(v) -> str:
    return str(v or "").strip()


def is_alias_row(row: dict) -> bool:
    fn = norm(row.get("First Name (Localized)", ""))
    ln = norm(row.get("Last Name (Localized)", ""))
    return bool(CHAPTER_RE.match(fn) or ln == "Yes")


# ─── Finding data class ───────────────────────────────────────────────────────
@dataclass
class Finding:
    stage:          str   # S1-S5 or LLM
    severity:       str   # critical | high | medium | low | info
    kind:           str   # snake_case tag
    detail:         str
    fix:            str = ""
    suggested_name: str = ""  # populated by cascade / cross-batch checks


SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}


def worst_severity(findings: list[Finding]) -> str:
    if not findings:
        return "pass"
    return min(findings, key=lambda f: SEV_ORDER.get(f.severity, 9)).severity


def to_human_format(findings: list[Finding], existing: str = "") -> str:
    """Render findings in the human verifier cell format, preserving existing text."""
    parts = [p.strip() for p in existing.split("|") if p.strip()] if existing else []
    for f in findings:
        tag = KIND_TAG.get(f.kind, f.kind.upper())
        sev = f.severity if f.severity in ("high", "medium", "low") else "medium"
        text = f"{tag}: {f.detail}"
        if f.fix:
            text += f" {f.fix}"
        if f.suggested_name:
            text += f" Suggested: '{f.suggested_name}'."
        text += f" (Severity: {sev})"
        if text not in parts:
            parts.append(text)
    return " | ".join(parts)


# ─── xlsx I/O helpers ─────────────────────────────────────────────────────────
# Map the many header spellings used across LS exports to one canonical key set,
# so all downstream logic can use stable names regardless of the source schema.
COLUMN_ALIASES = {
    "Type":                    ["type"],
    "ID":                      ["id"],
    "Original Name":           ["original name", "original_name", "original mention",
                                "original_mention"],
    "Localized Name":          ["localized name", "localised name", "localized_name",
                                "localised_name", "localized mention", "localised mention",
                                "localized_mention", "localised_mention"],
    "English Translated Name": ["english translated name", "english translated mention"],
    "First Name (Original)":   ["first name (original)", "first_name_original"],
    "Last Name (Original)":    ["last name (original)", "last_name_original"],
    "First Name (Localized)":  ["first name (localized)", "first name (localised)",
                                "first_name_localized", "first_name_localised"],
    "Last Name (Localized)":   ["last name (localized)", "last name (localised)",
                                "last_name_localized", "last_name_localised"],
    "Gender":                  ["gender"],
    "Canonical Name":          ["canonical name", "canonical_name"],
    "Chapter Numbers":         ["chapter numbers", "chapter_numbers"],
    "Is New":                  ["is new", "is_new"],
    "Localization Issues":     ["localization issues", "localisation issues"],
    "Cultural Status":         ["cultural status", "cultural_status"],
    "Localization Reason":     ["localization reason", "localisation reason"],
    "Description":             ["description", "notes"],
}


def _canon_key(header: str) -> Optional[str]:
    hl = header.strip().lower()
    for canon, aliases in COLUMN_ALIASES.items():
        if hl == canon.lower() or hl in aliases:
            return canon
    return None


def read_sheet(wb: openpyxl.Workbook, name) -> tuple[list[str], list[dict]]:
    """Read a sheet into a list of row dicts.

    Each row dict is keyed by BOTH the original header AND (when recognised) the
    canonical column name, so callers may use stable names like 'Original Name'
    even when the sheet header is 'original_name'.
    """
    ws = wb[name] if isinstance(name, str) else wb.worksheets[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    canon = [_canon_key(h) for h in headers]
    data = []
    for row in rows[1:]:
        d: dict = {}
        for i, h in enumerate(headers):
            val = str(row[i]) if i < len(row) and row[i] is not None else ""
            d[h] = val
            if canon[i] and canon[i] not in d:
                d[canon[i]] = val
        data.append(d)
    return headers, data


def find_col(headers: list[str], name: str) -> int:
    """Return 1-based column index, or -1 if not found (case-insensitive)."""
    nl = name.lower()
    for i, h in enumerate(headers):
        if h.lower() == nl or nl in h.lower():
            return i + 1
    return -1


def ensure_col(ws, headers: list[str], name: str) -> int:
    """Ensure column exists; return its 1-based index."""
    idx = find_col(headers, name)
    if idx < 0:
        idx = len(headers) + 1
        ws.cell(row=1, column=idx, value=name)
        headers.append(name)
    return idx


# ─── Stage 1: Master ingestion ────────────────────────────────────────────────
@dataclass
class Universe:
    characters: dict = field(default_factory=dict)  # orig_lower → {localized, description, type}
    entities:   dict = field(default_factory=dict)  # places, orgs, objects
    genre:      str = ""
    setting:    str = ""
    raw_notes:  list[str] = field(default_factory=list)


def load_master(path: str) -> Universe:
    """
    Parse the Master adaptation sheet (xlsx).

    Accepts flexible column layouts; looks for any column containing the keywords
    below.  Falls back to reading every cell as raw universe context if no
    structured columns are found.
    """
    u = Universe()
    if not path or not os.path.exists(path):
        return u

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdrs = [str(h or "").strip().lower() for h in rows[0]]

        orig_col = next((i for i, h in enumerate(hdrs)
                         if any(k in h for k in ["original", "source", "character name", "nom original"])), -1)
        loc_col  = next((i for i, h in enumerate(hdrs)
                         if any(k in h for k in ["locali", "translation", "adapted", "nom localisé"])), -1)
        desc_col = next((i for i, h in enumerate(hdrs)
                         if any(k in h for k in ["description", "notes", "context", "profile", "personnage"])), -1)
        type_col = next((i for i, h in enumerate(hdrs)
                         if any(k in h for k in ["type", "role", "category", "catégorie"])), -1)

        for row in rows[1:]:
            if not any(c for c in row if c):
                continue
            raw = " | ".join(str(c or "") for c in row if c)
            u.raw_notes.append(raw)

            if orig_col >= 0 and orig_col < len(row):
                orig = str(row[orig_col] or "").strip()
                loc  = str(row[loc_col]  or "").strip() if loc_col  >= 0 else ""
                desc = str(row[desc_col] or "").strip() if desc_col >= 0 else ""
                typ  = str(row[type_col] or "").strip().lower() if type_col >= 0 else ""
                if orig:
                    entry = {"localized": loc, "description": desc, "type": typ}
                    if any(k in typ for k in ("place", "location", "city", "lieu", "org")):
                        u.entities[orig.lower()] = entry
                    else:
                        u.characters[orig.lower()] = entry

    print(f"  Master loaded: {len(u.characters)} characters, "
          f"{len(u.entities)} entities, {len(u.raw_notes)} context lines.",
          file=sys.stderr)
    return u


# ─── Stage 2: Proposition review ─────────────────────────────────────────────
def review_proposition(row: dict, universe: Universe, src: str, tgt: str) -> list[Finding]:
    """
    Deterministic checks on one Localization Details main row:
    - Missing required fields (C0)
    - Name-part alignment (C1)
    - Source-language leak / status contradiction (C2)
    - Master-sheet cross-check (S2)
    """
    if is_alias_row(row):
        return []

    findings: list[Finding] = []
    orig   = norm(row.get("Original Name", ""))
    loc    = norm(row.get("Localized Name", ""))
    fn     = norm(row.get("First Name (Localized)", ""))
    ln     = norm(row.get("Last Name (Localized)", ""))
    status = norm(row.get("Cultural Status", "")).upper()
    reason = norm(row.get("Localization Reason", ""))

    # skip blank / spacer rows entirely
    if is_empty(orig):
        return []

    # C0: missing localisation
    if is_empty(loc):
        findings.append(Finding("S2", "critical", "missing_localised",
                                f"'{orig}' has no {lang_name(tgt)} localized name.",
                                "Add a localized name."))
        return findings  # further checks are meaningless without a localized name

    # C0: metadata completeness
    if is_empty(status):
        findings.append(Finding("S2", "high", "missing_cultural_status",
                                f"Cultural Status missing for '{orig}'.",
                                "Set a cultural status (LOCALIZED/KEPT/ADAPTED/…)."))
    elif status not in VALID_CULTURAL:
        findings.append(Finding("S2", "medium", "invalid_cultural_status",
                                f"Unrecognised status '{status}' for '{orig}'.",
                                f"Use one of: {', '.join(sorted(VALID_CULTURAL))}."))

    if is_empty(reason):
        findings.append(Finding("S2", "medium", "missing_localisation_reason",
                                f"Localization Reason missing for '{orig}'.",
                                "Explain the localization decision."))

    # C1: first + last name match the full localized name
    if not is_empty(fn) and not is_empty(ln):
        expected = f"{fn} {ln}"
        if expected.lower() != loc.lower():
            findings.append(Finding("S2", "medium", "name_part_mismatch",
                                    f"'{fn} {ln}' ≠ '{loc}' for '{orig}'.",
                                    f"Align name parts. Suggested full name: '{expected}'.",
                                    suggested_name=expected))

    # C2: source-language leak
    leak_re = SOURCE_LEAK.get(src.split("-")[0])
    if leak_re:
        m = leak_re.search(loc)
        if m:
            findings.append(Finding("S2", "high", "source_lang_leak",
                                    f"{lang_name(src)} word '{m.group()}' found in localised name '{loc}'.",
                                    f"Replace with {lang_name(tgt)} equivalent."))

    # C2: name unchanged but status doesn't say so
    if loc.lower() == orig.lower() and status not in {"KEPT", "ORIGINAL", "UNIVERSAL", ""}:
        findings.append(Finding("S2", "high", "source_name_not_localised",
                                f"'{loc}' matches the source name '{orig}' but Cultural Status is '{status}'.",
                                "Either localize the name or set status to KEPT/ORIGINAL."))

    # C2: status says KEPT/ORIGINAL but name was changed
    if status in {"KEPT", "ORIGINAL"} and loc.lower() != orig.lower():
        findings.append(Finding("S2", "high", "status_mismatch",
                                f"Status is '{status}' but name changed: '{orig}' → '{loc}'.",
                                "Fix the status or revert the name."))

    # C3: casing
    if LOWER_START.match(loc):
        findings.append(Finding("S2", "low", "casing",
                                f"'{loc}' starts with a lowercase letter.",
                                "Capitalise the first letter."))

    # S2: Master cross-check — if Master has a different localized form
    char_info = universe.characters.get(orig.lower(), {})
    master_loc = char_info.get("localized", "")
    if master_loc and not is_empty(master_loc) and master_loc.lower() != loc.lower():
        findings.append(Finding("S2", "medium", "master_mismatch",
                                f"Master sheet says '{orig}' → '{master_loc}', but LS has '{loc}'.",
                                f"Consider adopting '{master_loc}' from the Master.",
                                suggested_name=master_loc))

    return findings


# ─── Name collision check (new Stage 2b) ─────────────────────────────────────
DE_KINSHIP_RE = re.compile(
    r"\b(onkel|tante|bruder|schwester|schwägerin|großvater|großmutter|oma|opa|schatz)\b",
    re.IGNORECASE,
)


def check_name_collisions(
    ld_rows: list[dict],
    src: str,
    tgt: str,
    mm_rows: Optional[list[dict]] = None,
) -> list[tuple[int, list[Finding]]]:
    """
    Flag localized names shared by multiple DISTINCT characters.

    Crucially, the collision pool spans BOTH the Localization Details rows and
    every localized form in the Mention Mappings tab — because a clashing name
    often lives only in the mention map (e.g. another character is *referred to*
    as 'Pierre'). Characters are de-duplicated by their canonical original name,
    so a character colliding with its own alias is not counted.

    SAME_FIRST_NAME_COLLISION — same localized first name used by 2+ characters.
    SAME_LAST_NAME_COLLISION  — same localized last name across different families.
    CULTURAL_CONTEXT_INAPPROPRIATE — source-language kinship term retained.
    """
    # localized-name-token -> set of distinct character identities (orig name)
    fn_chars: dict[str, set[str]] = defaultdict(set)
    ln_chars: dict[str, set[str]] = defaultdict(set)

    def add_first(tok: str, ident: str):
        if tok and not is_empty(tok) and len(tok) > 2:
            fn_chars[_apos(tok)].add(ident.lower())

    def add_last(tok: str, ident: str):
        if tok and not is_empty(tok) and tok.lower() != "yes" and not CHAPTER_RE.match(tok):
            ln_chars[_apos(tok)].add(ident.lower())

    # LD main rows
    for row in ld_rows:
        if is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        if not orig:
            continue
        add_first(norm(row.get("First Name (Localized)", "")), orig)
        add_last(norm(row.get("Last Name (Localized)", "")), orig)

    # Mention Mappings — localized mention forms, identified by Canonical Name
    if mm_rows:
        for row in mm_rows:
            if norm(row.get("Type", "")).lower() not in ("", "character"):
                continue
            cn = norm(row.get("Canonical Name", ""))
            lm = norm(row.get("Localized Name", ""))
            if not cn or is_empty(lm):
                continue
            if any(k in lm.lower() for k in KINSHIP):
                continue
            core = _strip_title(lm)
            toks = _PARTICLE_RE.sub("", core).strip().split()
            if not toks:
                continue
            add_first(toks[0], cn)
            if len(toks) > 1:
                add_last(toks[-1], cn)

    row_findings: list[tuple[int, list[Finding]]] = []

    for i, row in enumerate(ld_rows):
        if is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        loc  = norm(row.get("Localized Name", ""))
        fn   = norm(row.get("First Name (Localized)", ""))
        ln   = norm(row.get("Last Name (Localized)", ""))
        if not orig or is_empty(loc):
            continue
        ident = orig.lower()
        orig_surname = orig.split()[-1].lower() if " " in orig else orig.lower()

        findings: list[Finding] = []

        # SAME_FIRST_NAME_COLLISION — shared by 2+ distinct characters
        if fn and not is_empty(fn):
            chars = fn_chars.get(_apos(fn), set())
            others = sorted(chars - {ident})
            if others:
                findings.append(Finding(
                    "S2b", "medium", "same_first_name_collision",
                    f"Character '{orig}' shares localized first name '{fn}' with: "
                    f"{', '.join(others)}. Listener confusion risk. "
                    f"+ affected mentions: {fn}",
                    "Consider a different localized first name to avoid ambiguity.",
                ))

        # SAME_LAST_NAME_COLLISION — shared across different original families
        if ln and not is_empty(ln) and ln.lower() != "yes" and not CHAPTER_RE.match(ln):
            chars = ln_chars.get(_apos(ln), set())
            other_families = {
                (o.split()[-1].lower() if " " in o else o.lower())
                for o in chars
            } - {orig_surname}
            if other_families:
                others = sorted(o for o in chars
                                if (o.split()[-1].lower() if " " in o else o.lower())
                                != orig_surname)
                findings.append(Finding(
                    "S2b", "medium", "same_last_name_collision",
                    f"Character '{orig}' shares localized last name '{ln}' with character(s) "
                    f"from a different family: {', '.join(others)}. May create confusion. "
                    f"+ affected mentions: {ln}",
                    "Consider a different localized last name.",
                ))

        # CULTURAL_CONTEXT_INAPPROPRIATE (German kinship retained in target)
        if src.startswith("de") and DE_KINSHIP_RE.search(loc):
            m = DE_KINSHIP_RE.search(loc)
            findings.append(Finding(
                "S2b", "high", "cultural_context_inappropriate",
                f"German kinship term '{m.group()}' retained in localized name '{loc}' "
                f"for '{orig}'. Should be adapted to {lang_name(tgt)} equivalent. "
                f"+ affected mentions: {orig}",
                f"Replace '{m.group()}' with its {lang_name(tgt)} equivalent.",
            ))

        if findings:
            row_findings.append((i, findings))

    return row_findings


# ─── Lever 2: Entity / family consistency ────────────────────────────────────
_FAMILY_ORIG_RE = re.compile(
    r"^([a-zà-öø-ÿ]+)[\s\-](famili|family|dynasti|dynasty|clan|familie)\b", re.IGNORECASE)
# pull the localized surname out of "la famille X" / "la dynastie X" / "le clan X"
_FAMILY_LOC_RE = re.compile(
    r"(?:famille|dynastie|clan|groupe)\s+(?:de\s+|des\s+)?([A-ZÀ-Ö][\wà-öø-ÿ’'\-]+)")


def build_family_map(ld_rows: list[dict]) -> dict[str, str]:
    """
    Extract {original_surname_lower: localized_surname} from family entities such
    as 'kaiser-familie' -> 'la famille Montclair'  =>  {'kaiser': 'Montclair'}.

    A surname maps only if we can cleanly extract a capitalised target surname.
    """
    fam: dict[str, str] = {}
    for row in ld_rows:
        if norm(row.get("Type", "")).lower() != "entity":
            continue
        orig = norm(row.get("Original Name", "")).lower()
        loc  = norm(row.get("Localized Name", ""))
        m = _FAMILY_ORIG_RE.match(orig)
        if not m:
            continue
        surname = m.group(1).strip()
        lm = _FAMILY_LOC_RE.search(loc)
        if not lm:
            continue
        loc_surname = lm.group(1).strip()
        # first clean mapping wins (the canonical "<surname>-familie" entity)
        fam.setdefault(surname, loc_surname)
    return fam


def _apos(s: str) -> str:
    """Normalise curly/straight apostrophes for comparison."""
    return s.replace("’", "'").replace("ʼ", "'").lower()


def check_entity_consistency(
    ld_rows: list[dict],
) -> list[tuple[int, list[Finding]]]:
    """
    Lever 2 — ENTITY_INCONSISTENCY:
    A character whose ORIGINAL surname belongs to a known family entity must use
    that family's localized surname. Flags e.g. 'jacob kaiser' localized with
    surname 'Barnier' when the Kaiser family entity localizes to 'Montclair'.
    """
    fam = build_family_map(ld_rows)
    row_findings: list[tuple[int, list[Finding]]] = []
    if not fam:
        return row_findings

    for i, row in enumerate(ld_rows):
        if norm(row.get("Type", "")).lower() != "character" or is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        ln   = norm(row.get("Last Name (Localized)", ""))
        if " " not in orig or is_empty(ln) or ln.lower() == "yes":
            continue
        osurname = orig.split()[-1].lower()
        expected = fam.get(osurname)
        if not expected:
            continue
        # accept compound/hyphenated surnames that contain the family surname
        # e.g. 'Féretvin-Barnier' satisfies an expected 'Barnier'
        ln_tokens = {_apos(t) for t in re.split(r"[\s\-]+", ln) if t}
        if _apos(expected) not in ln_tokens and _apos(ln) != _apos(expected):
            row_findings.append((i, [Finding(
                "S2c", "high", "entity_inconsistency",
                f"Character '{orig}' has localized surname '{ln}', but the '{osurname}' "
                f"family entity is localized as '{expected}'. Surname should be '{expected}'. "
                f"+ affected mentions: {orig}",
                f"Use the family surname '{expected}'.",
                suggested_name=expected,
            )]))
    return row_findings


# ─── Lever 1: Mention-map internal consistency ───────────────────────────────
_TITLE_RE = re.compile(
    r"^(monsieur|madame|mademoiselle|m\.|mme\.?|mlle\.?|dr\.?|docteur|professeur|"
    r"prof\.?|maître|me\.?|sir|lord|lady)\s+", re.IGNORECASE)


def _strip_title(s: str) -> str:
    return _TITLE_RE.sub("", s).strip()


# leading French articles / possessive particles to strip before comparing names
_PARTICLE_RE = re.compile(r"^(d'|l'|de\s+la\s+|de\s+|du\s+|des\s+|le\s+|la\s+|les\s+)",
                          re.IGNORECASE)


def _name_first_token(s: str) -> str:
    """Strip title + leading article/particle, return the first name token."""
    core = _strip_title(s)
    core = _PARTICLE_RE.sub("", core).strip()
    core = _PARTICLE_RE.sub("", core).strip()  # handle "de l'" style doubles
    toks = core.split()
    return toks[0] if toks else ""


def check_mention_consistency(
    mm_rows: list[dict],
    ld_rows: list[dict],
) -> list[tuple[int, list[Finding]]]:
    """
    Lever 1 — group character Mention Mappings by Canonical Name and verify every
    mention's localized FIRST name is stable across the character's mentions.

    Conservative: only CHARACTER-type mentions whose ORIGINAL form actually uses
    the character's first name are judged (title+surname-only forms are skipped,
    since they carry no first name to check), and French articles/possessives are
    stripped before comparison to avoid false 'd'Aurore' ≠ 'Aurore' flags.

    Produces MENTION_MAP_INCONSISTENCY and SOURCE_NAME_NOT_LOCALISED.
    """
    # group CHARACTER mention rows by canonical name
    groups: dict[str, list[int]] = defaultdict(list)
    for i, row in enumerate(mm_rows):
        if norm(row.get("Type", "")).lower() not in ("", "character"):
            continue
        cn = norm(row.get("Canonical Name", "")).lower()
        if cn:
            groups[cn].append(i)

    row_findings: list[tuple[int, list[Finding]]] = []

    for cn, idxs in groups.items():
        orig_first = cn.split()[0] if cn else ""
        if not orig_first or len(orig_first) < 3:
            continue

        # ── collect this character's localized first-name tokens, only from
        #    mentions that actually USE the source first name ────────────────
        # token -> {"count": n, "rows": [...], "src_leak": bool}
        token_rows: dict[str, list[int]] = defaultdict(list)
        token_src_leak: dict[str, bool] = {}
        for i in idxs:
            om = norm(mm_rows[i].get("Original Name", ""))
            lm = norm(mm_rows[i].get("Localized Name", ""))
            if is_empty(om) or is_empty(lm):
                continue
            oml = om.lower()
            om_head = oml.split()[0].rstrip("s").rstrip("'")
            if not om_head.startswith(orig_first[:4]):
                continue
            if any(k in oml for k in KINSHIP) or any(k in lm.lower() for k in KINSHIP):
                continue
            tok = _name_first_token(lm)
            if not tok or len(tok) < 3:
                continue
            key = _apos(tok)
            token_rows[key].append(i)
            # is this token just the untranslated source name?
            token_src_leak[key] = (key == orig_first) or (orig_first in _apos(lm).split())

        if len(token_rows) < 2:
            continue  # internally consistent → nothing to flag

        # the majority token is the canonical localized first name; minorities
        # are the inconsistencies to flag
        majority = max(token_rows, key=lambda k: len(token_rows[k]))
        canon_disp = majority.capitalize()

        for tok, rows in token_rows.items():
            if tok == majority:
                continue
            for i in rows:
                om = norm(mm_rows[i].get("Original Name", ""))
                lm = norm(mm_rows[i].get("Localized Name", ""))
                if token_src_leak.get(tok):
                    row_findings.append((i, [Finding(
                        "S1", "medium", "source_name_not_localised",
                        f"Mention '{om}' keeps the source name in its localization "
                        f"'{lm}', while '{cn}' is elsewhere localized as '{canon_disp}'. "
                        f"+ affected mentions: {om}",
                        f"Localize to '{canon_disp}'.",
                        suggested_name=canon_disp,
                    )]))
                else:
                    row_findings.append((i, [Finding(
                        "S1", "high", "mention_map_inconsistency",
                        f"Character '{cn}' is localized inconsistently: mention '{om}' "
                        f"uses '{lm}' but other mentions use '{canon_disp}'. "
                        f"+ affected mentions: {om}",
                        f"Use '{canon_disp}' consistently.",
                        suggested_name=canon_disp,
                    )]))

    return row_findings


# ─── Stage 4: Uniformly flat rule ────────────────────────────────────────────
def check_flat_consistency(
    ld_rows: list[dict],
) -> tuple[dict[str, str], list[tuple[int, list[Finding]]]]:
    """
    Ensure that every occurrence of the same source name maps to the same
    localized name within the current batch.

    Returns:
        canonical: orig_lower → most-common localized form
        row_findings: (row_index, findings) for deviating rows
    """
    tally: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in ld_rows:
        if is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        loc  = norm(row.get("Localized Name", ""))
        if not is_empty(orig) and not is_empty(loc):
            tally[orig.lower()][loc] += 1

    canonical: dict[str, str] = {k: max(v, key=v.get) for k, v in tally.items()}

    row_findings: list[tuple[int, list[Finding]]] = []
    for i, row in enumerate(ld_rows):
        if is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        loc  = norm(row.get("Localized Name", ""))
        if is_empty(orig) or is_empty(loc):
            continue
        canon = canonical.get(orig.lower(), "")
        if canon and loc != canon:
            row_findings.append((i, [Finding(
                "S4", "high", "cross_character_inconsistency",
                f"'{orig}' is localized as '{loc}' here but '{canon}' elsewhere in this batch.",
                f"Use '{canon}' consistently throughout.",
                suggested_name=canon,
            )]))

    return canonical, row_findings


# ─── Stage 3: Mention mapping cascade ────────────────────────────────────────
def check_mention_cascade(
    mm_rows: list[dict],
    canonical: dict[str, str],
) -> list[tuple[int, list[Finding]]]:
    """
    Stage 3: verify that every Mention Mapping entry uses the canonical localized
    form decided in the Localization Details tab.

    Skips kinship terms (they don't follow character-name patterns).
    """
    row_findings: list[tuple[int, list[Finding]]] = []

    for i, row in enumerate(mm_rows):
        orig_mention = norm(row.get("Original Mention") or row.get("Original Name", ""))
        loc_mention  = norm(row.get("Localized Mention") or row.get("Localized Name", ""))
        canon_name   = norm(row.get("Canonical Name", ""))

        if is_empty(orig_mention) or is_empty(loc_mention):
            continue

        # Skip kinship terms
        if any(k in orig_mention.lower() for k in KINSHIP):
            continue
        if any(k in loc_mention.lower() for k in KINSHIP):
            continue

        # Find canonical translation for this character
        lookup_key = canon_name.lower() or orig_mention.lower()
        canon_loc  = canonical.get(lookup_key, "")

        # Partial-match fallback (canonical name may be a substring)
        if not canon_loc:
            for ck, cv in canonical.items():
                if ck and (ck in lookup_key or lookup_key in ck):
                    canon_loc = cv
                    break

        if not canon_loc:
            continue

        # Compare first-name tokens to allow inflections/particles
        canon_first = canon_loc.split()[0].lower()
        loc_first   = loc_mention.split()[0].lower()

        if canon_first and loc_first and canon_first != loc_first:
            # Only flag when the canonical name appears in the original mention
            if canon_name.lower() and canon_name.lower() not in orig_mention.lower():
                continue
            row_findings.append((i, [Finding(
                "S3", "high", "mention_map_inconsistency",
                f"'{orig_mention}' → '{loc_mention}' "
                f"but canonical localization is '{canon_loc}'.",
                f"Update mention to use '{canon_loc}' consistently.",
                suggested_name=canon_loc,
            )]))

    return row_findings


# ─── Stage 5: Cross-batch consistency ────────────────────────────────────────
def load_batch_precedent(path: str) -> dict[str, str]:
    """
    Read a previous-batch xlsx and return {orig_lower: localized_name}.
    Looks for the Localization Details tab first.
    """
    precedent: dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # Prefer the Localization Details tab
        tab = next(
            (n for n in wb.sheetnames if "detail" in n.lower() or "locali" in n.lower()),
            wb.sheetnames[0],
        )
        _, rows = read_sheet(wb, tab)
        for row in rows:
            orig = norm(row.get("Original Name", ""))
            loc  = norm(row.get("Localized Name", ""))
            if not is_empty(orig) and not is_empty(loc) and not is_alias_row(row):
                precedent[orig.lower()] = loc
        print(f"  Loaded '{path}': {len(precedent)} precedents.", file=sys.stderr)
    except Exception as e:
        print(f"  WARNING: could not read previous batch '{path}': {e}", file=sys.stderr)
    return precedent


def check_cross_batch(
    ld_rows: list[dict],
    precedents: list[dict[str, str]],
) -> list[tuple[int, list[Finding]]]:
    """
    Stage 5: flag any row where the current localization contradicts a prior batch.

    Example: if 'Shirley' was 'Agathe' in Eps 1-100, it must not become
    'Charlène' in Eps 101-200.
    """
    row_findings: list[tuple[int, list[Finding]]] = []

    # Merge all precedent batches; earlier batches win
    merged: dict[str, str] = {}
    for batch in precedents:
        for k, v in batch.items():
            merged.setdefault(k, v)

    for i, row in enumerate(ld_rows):
        if is_alias_row(row):
            continue
        orig = norm(row.get("Original Name", ""))
        loc  = norm(row.get("Localized Name", ""))
        if is_empty(orig) or is_empty(loc):
            continue
        prior = merged.get(orig.lower(), "")
        if prior and prior.lower() != loc.lower():
            row_findings.append((i, [Finding(
                "S5", "critical", "cross_batch_inconsistency",
                f"'{orig}' was localized as '{prior}' in a previous batch "
                f"but this batch uses '{loc}'.",
                f"Revert to '{prior}' to maintain cross-batch consistency.",
                suggested_name=prior,
            )]))

    return row_findings


# ─── LLM proposition review (optional) ───────────────────────────────────────
SYS_PROMPT = (
    "You are an expert localization quality reviewer for fiction audiobooks. "
    "You review whether character and entity names have been appropriately localized "
    "from {src} into {tgt}. A good localization: (1) sounds natural and fluent in {tgt}, "
    "(2) fits the character's personality, origin, and role, (3) uses the right register, "
    "(4) avoids cultural inappropriateness, (5) is consistent with the show's genre. "
    "Be precise and conservative — only flag genuine problems."
)

USR_PROMPT = (
    "Review this localization proposition:\n\n"
    "Original ({src}) name : {orig}\n"
    "Localized ({tgt}) name: {loc}\n"
    "Cultural status       : {status}\n"
    "Stated reason         : {reason}\n"
    "Character description : {desc}\n"
    "Show universe context : {universe}\n\n"
    "Respond with STRICT JSON only:\n"
    '{{"verdict":"pass|warn|fail",'
    '"issues":[{{"severity":"high|medium|low","kind":"snake_tag",'
    '"detail":"one sentence","fix":"one sentence","suggested_name":""}}],'
    '"overall_comment":"<=20 words"}}\n'
    "If the localization is good, return verdict \"pass\" and an empty issues list."
)


@dataclass
class LLMConfig:
    api_key:     str
    base_url:    str
    model:       str
    max_retries: int = 4


def build_client(cfg: LLMConfig):
    from openai import OpenAI
    return OpenAI(api_key=cfg.api_key, base_url=cfg.base_url, max_retries=0)


def llm_review_row(client, cfg: LLMConfig, row: dict, universe: Universe,
                   src: str, tgt: str) -> list[Finding]:
    orig = norm(row.get("Original Name", ""))
    loc  = norm(row.get("Localized Name", ""))
    if is_empty(orig) or is_empty(loc) or is_alias_row(row):
        return []

    char_info    = universe.characters.get(orig.lower(), {})
    universe_ctx = char_info.get("description", "") or " | ".join(universe.raw_notes[:5])

    prompt = USR_PROMPT.format(
        src=lang_name(src), tgt=lang_name(tgt),
        orig=orig, loc=loc,
        status=norm(row.get("Cultural Status", "")) or "(none)",
        reason=norm(row.get("Localization Reason", ""))[:500] or "(none)",
        desc=norm(row.get("Description", ""))[:500] or "(none)",
        universe=(universe_ctx or "(no master context)")[:800],
    )
    messages = [
        {"role": "system", "content": SYS_PROMPT.format(src=lang_name(src), tgt=lang_name(tgt))},
        {"role": "user",   "content": prompt},
    ]
    last_err = None
    for attempt in range(cfg.max_retries):
        try:
            resp = client.chat.completions.create(
                model=cfg.model, messages=messages, temperature=0.0,
                response_format={"type": "json_object"},
            )
            data = json.loads(resp.choices[0].message.content or "{}")
            return [
                Finding("LLM",
                        (it.get("severity") or "medium").lower(),
                        it.get("kind") or "llm_proposition",
                        it.get("detail") or "",
                        it.get("fix") or "",
                        suggested_name=it.get("suggested_name") or "")
                for it in (data.get("issues") or [])
            ]
        except (json.JSONDecodeError, Exception) as e:
            last_err = e
            time.sleep(2 ** attempt)
    return [Finding("LLM", "info", "llm_error",
                    f"LLM call failed: {type(last_err).__name__}: {str(last_err)[:100]}")]


# ─── Write verified output ────────────────────────────────────────────────────
ISSUES_COL = "Localization Issues"
NOTES_COL  = "Verification Notes"


def write_ld_findings(ws, headers: list[str],
                      findings_map: dict[int, list[Finding]],
                      existing_issues: list[str]) -> None:
    """Write findings into the Localization Issues column of the LD sheet."""
    col = ensure_col(ws, headers, ISSUES_COL)
    for row_i, findings in findings_map.items():
        if not findings:
            continue
        ws_row = row_i + 2  # +1 for header, +1 for 1-based
        cell = ws.cell(row=ws_row, column=col)
        existing = existing_issues[row_i] if row_i < len(existing_issues) else ""
        cell.value = to_human_format(findings, existing)
        fill = SEV_FILL.get(worst_severity(findings))
        if fill:
            cell.fill = fill


def write_mm_findings(ws, headers: list[str],
                      findings_map: dict[int, list[Finding]]) -> None:
    """Write findings into the Verification Notes column of the MM sheet."""
    col = ensure_col(ws, headers, NOTES_COL)
    for row_i, findings in findings_map.items():
        if not findings:
            continue
        ws_row = row_i + 2
        cell = ws.cell(row=ws_row, column=col)
        cell.value = to_human_format(findings, "")
        fill = SEV_FILL.get(worst_severity(findings))
        if fill:
            cell.fill = fill


# ─── Main pipeline ────────────────────────────────────────────────────────────
def run(args) -> int:
    src, tgt = args.source_lang, args.target_lang
    print(f"\n{'='*60}", file=sys.stderr)
    print(f"  Localization Pipeline: {lang_name(src)} → {lang_name(tgt)}", file=sys.stderr)
    print(f"{'='*60}\n", file=sys.stderr)

    # ── Stage 1: Master ingestion ─────────────────────────────────────────────
    print("Stage 1: Master ingestion", file=sys.stderr)
    if args.master:
        universe = load_master(args.master)
    else:
        universe = Universe()
        print("  (no --master provided; universe context will be absent)", file=sys.stderr)

    # ── Load input xlsx ───────────────────────────────────────────────────────
    print(f"\nLoading: {args.input}", file=sys.stderr)
    wb_in = openpyxl.load_workbook(args.input, data_only=True)

    ld_name = next((n for n in wb_in.sheetnames if "detail" in n.lower()), None) \
              or wb_in.sheetnames[0]
    mm_name = next((n for n in wb_in.sheetnames if "mention" in n.lower()), None) \
              or (wb_in.sheetnames[1] if len(wb_in.sheetnames) > 1 else None)

    ld_headers, ld_rows = read_sheet(wb_in, ld_name)
    mm_headers, mm_rows = read_sheet(wb_in, mm_name) if mm_name else ([], [])

    print(f"  '{ld_name}': {len(ld_rows)} rows", file=sys.stderr)
    if mm_name:
        print(f"  '{mm_name}': {len(mm_rows)} rows", file=sys.stderr)

    existing_issues = [norm(r.get(ISSUES_COL, "")) for r in ld_rows]

    # ── Set up LLM ────────────────────────────────────────────────────────────
    client = cfg = None
    if not args.dry_run:
        api_key  = os.environ.get("ARGUS_API_KEY")
        base_url = os.environ.get("ARGUS_BASE_URL")
        model    = args.model or os.environ.get("ARGUS_MODEL", "claude-opus-4.8")
        if not api_key or not base_url:
            print("ERROR: set ARGUS_API_KEY and ARGUS_BASE_URL, or use --dry-run.",
                  file=sys.stderr)
            return 2
        cfg    = LLMConfig(api_key=api_key, base_url=base_url, model=model)
        client = build_client(cfg)
        print(f"\n  Argus: {base_url}  model={model}", file=sys.stderr)

    # ── Stage 2: Proposition review ───────────────────────────────────────────
    print("\nStage 2: Proposition review (deterministic)", file=sys.stderr)
    ld_findings: dict[int, list[Finding]] = defaultdict(list)

    for i, row in enumerate(ld_rows):
        fs = review_proposition(row, universe, src, tgt)
        ld_findings[i].extend(fs)

    s2_count = sum(1 for fs in ld_findings.values() if fs)
    print(f"  {s2_count} rows with issues.", file=sys.stderr)

    # ── Stage 2b: Name collision checks ──────────────────────────────────────
    print("\nStage 2b: Name collision checks", file=sys.stderr)
    coll_findings = check_name_collisions(ld_rows, src, tgt, mm_rows)
    for i, fs in coll_findings:
        ld_findings[i].extend(fs)
    fn_c = sum(1 for _, fs in coll_findings for f in fs if "first_name" in f.kind)
    ln_c = sum(1 for _, fs in coll_findings for f in fs if "last_name" in f.kind)
    cc_c = sum(1 for _, fs in coll_findings for f in fs if "cultural" in f.kind)
    print(f"  {fn_c} first-name collisions, {ln_c} last-name collisions, "
          f"{cc_c} cultural context issues.", file=sys.stderr)

    # ── Lever 2: Entity / family consistency ─────────────────────────────────
    print("\nLever 2: Entity / family consistency", file=sys.stderr)
    fam = build_family_map(ld_rows)
    ent_findings = check_entity_consistency(ld_rows)
    for i, fs in ent_findings:
        ld_findings[i].extend(fs)
    print(f"  {len(fam)} family entities mapped, "
          f"{len(ent_findings)} surname inconsistencies flagged.", file=sys.stderr)

    # LLM proposition review (optional)
    if not args.dry_run and client and cfg:
        main_idx = [i for i, r in enumerate(ld_rows)
                    if not is_alias_row(r) and not is_empty(r.get("Localized Name", ""))]
        if args.limit:
            main_idx = main_idx[:args.limit]
        print(f"  LLM reviewing {len(main_idx)} main rows...", file=sys.stderr)
        done = 0
        with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
            fut = {ex.submit(llm_review_row, client, cfg, ld_rows[i], universe, src, tgt): i
                   for i in main_idx}
            for f_obj in as_completed(fut):
                i = fut[f_obj]
                ld_findings[i].extend(f_obj.result())
                done += 1
                if done % 10 == 0 or done == len(main_idx):
                    print(f"    …{done}/{len(main_idx)} LLM-reviewed", file=sys.stderr)

    # ── Stage 4: Uniformly flat rule ─────────────────────────────────────────
    print("\nStage 4: Uniformly flat rule", file=sys.stderr)
    canonical, flat_findings = check_flat_consistency(ld_rows)
    for i, fs in flat_findings:
        ld_findings[i].extend(fs)
    print(f"  {len(flat_findings)} inconsistencies.  Canonical map: {len(canonical)} entries.",
          file=sys.stderr)

    # ── Lever 1 + Stage 3: Mention-map consistency ───────────────────────────
    print("\nLever 1: Mention-map consistency", file=sys.stderr)
    mm_findings: dict[int, list[Finding]] = defaultdict(list)
    if mm_rows:
        mm_row_findings = check_mention_consistency(mm_rows, ld_rows)
        for i, fs in mm_row_findings:
            mm_findings[i].extend(fs)
        # legacy canonical cascade (catches cases the grouping misses)
        for i, fs in check_mention_cascade(mm_rows, canonical):
            mm_findings[i].extend(fs)
        mmi = sum(1 for _, fs in mm_row_findings for f in fs if "mention_map" in f.kind)
        snl = sum(1 for _, fs in mm_row_findings for f in fs if "source_name" in f.kind)
        print(f"  {mmi} mention inconsistencies, {snl} un-localized source names.",
              file=sys.stderr)
    else:
        print("  (no Mention Mappings tab found)", file=sys.stderr)

    # ── Stage 5: Cross-batch consistency ─────────────────────────────────────
    print("\nStage 5: Cross-batch consistency", file=sys.stderr)
    if args.previous:
        precedents = [load_batch_precedent(p) for p in args.previous]
        batch_findings = check_cross_batch(ld_rows, precedents)
        for i, fs in batch_findings:
            ld_findings[i].extend(fs)
        print(f"  {len(batch_findings)} cross-batch conflicts.", file=sys.stderr)
    else:
        print("  (no --previous batches; cross-batch check skipped)", file=sys.stderr)

    # ── Write output xlsx ─────────────────────────────────────────────────────
    print(f"\nWriting: {args.output}", file=sys.stderr)
    wb_out = openpyxl.load_workbook(args.input)

    ld_ws_out = wb_out[ld_name]
    write_ld_findings(ld_ws_out, list(ld_headers), ld_findings, existing_issues)

    if mm_name and mm_name in wb_out.sheetnames:
        mm_ws_out = wb_out[mm_name]
        write_mm_findings(mm_ws_out, list(mm_headers), mm_findings)

    wb_out.save(args.output)

    # ── Summary ───────────────────────────────────────────────────────────────
    total_ld  = sum(len(fs) for fs in ld_findings.values())
    total_mm  = sum(len(fs) for fs in mm_findings.values())
    critical  = sum(1 for fs in ld_findings.values() for f in fs if f.severity == "critical")
    high      = sum(1 for fs in ld_findings.values() for f in fs if f.severity == "high")
    medium    = sum(1 for fs in ld_findings.values() for f in fs if f.severity == "medium")

    print(f"\n{'='*60}", file=sys.stderr)
    print("  PIPELINE COMPLETE", file=sys.stderr)
    print(f"  LD findings : {total_ld} across "
          f"{sum(1 for v in ld_findings.values() if v)} rows", file=sys.stderr)
    print(f"  MM findings : {total_mm} across "
          f"{sum(1 for v in mm_findings.values() if v)} rows", file=sys.stderr)
    print(f"  Severity    : {critical} CRITICAL | {high} HIGH | {medium} MEDIUM",
          file=sys.stderr)
    print(f"  Output      : {args.output}", file=sys.stderr)
    if critical:
        print(f"\n  ⚠  {critical} CRITICAL issues require immediate attention.",
              file=sys.stderr)
    print(f"{'='*60}\n", file=sys.stderr)
    return 0


# ─── CLI ──────────────────────────────────────────────────────────────────────
def main() -> int:
    p = argparse.ArgumentParser(
        description="End-to-end localization pipeline (Stages 1-5)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--input",    required=True,
                   help="Source xlsx (needs Localization Details + Mention Mappings tabs)")
    p.add_argument("--output",   required=True,
                   help="Output xlsx with all verifications filled in")
    p.add_argument("--master",   default=None,
                   help="Master adaptation sheet xlsx (character bible / show universe)")
    p.add_argument("--previous", nargs="*", default=[],
                   help="Previous-batch xlsx files for cross-batch consistency (Stage 5). "
                        "Pass multiple files: --previous batch1.xlsx batch2.xlsx")
    p.add_argument("--source-lang",  default="de")
    p.add_argument("--target-lang",  default="fr")
    p.add_argument("--model",        default=None,
                   help="Override ARGUS_MODEL")
    p.add_argument("--concurrency",  type=int, default=4,
                   help="Parallel LLM calls (default 4)")
    p.add_argument("--limit",        type=int, default=0,
                   help="Cap on main rows sent to LLM (0 = all)")
    p.add_argument("--dry-run",      action="store_true",
                   help="Deterministic rule checks only — no LLM, no network, no key needed")
    return run(p.parse_args())


if __name__ == "__main__":
    sys.exit(main())
