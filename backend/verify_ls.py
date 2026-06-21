#!/usr/bin/env python3
"""
Localization-sheet verifier — importable engine + CLI.

  Layer 0  Shared corrections (learned from human edits)    → LEARNED / ALIAS_LEARNED
  Layer 1  Deterministic name-part matching                 → VERIFIED / MISMATCH / MISSING / SOURCE_LEAK
  Layer 2  Madeye LLM, one batched call per character        → LLM_OK / LLM_FIX

Story canon (from canon.pocketfm.ai) is authoritative: when the LLM detects a
conflict between the sheet and the canon (family, surname, alias), it returns
CANON_CONFLICT — flagged orange for human review, never auto-applied.

Usage (CLI):
    python verify_ls.py input.xlsx -o output.xlsx
    python verify_ls.py input.xlsx -o output.xlsx --llm

Env (for --llm):
    MADEYE_API_KEY     required
    MADEYE_BASE_URL    required
    MADEYE_MODEL       optional  (default claude-opus-4-8)
    MADEYE_USER_EMAIL  required  (Madeye needs metadata.user_email)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─── normalisation ─────────────────────────────────────────────────────────────
def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("œ", "oe").replace("æ", "ae")
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


SAFE_TITLES = {"mr": "Monsieur", "mister": "Monsieur", "mrs": "Madame",
               "miss": "Mademoiselle", "herr": "Monsieur", "frau": "Madame"}
CONNECTORS = {"de", "du", "des", "von", "der", "die", "das", "the", "of", "and",
              "et", "und", "le", "la", "les", "d", "l", "zu", "im", "den", "dem", "van"}

FILLS = {
    "VERIFIED":      "C6EFCE",
    "LEARNED":       "C6EFCE",
    "LLM_OK":        "C6EFCE",
    "ALIAS_LEARNED": "FFEB9C",
    "INCONSISTENT":  "FFEB9C",
    "MISMATCH":      "FCE4D6",
    "LLM_FIX":       "FCE4D6",
    "CANON_CONFLICT":"FCE4D6",
    "MISSING":       "FFC7CE",
    "SOURCE_LEAK":   "FFC7CE",
    "NAME_CHANGED":  "D9D2E9",
    "NEEDS_REVIEW":  "FFFFFF",
}


# ─── column helpers ────────────────────────────────────────────────────────────
def header_index(ws):
    return {norm(cell.value): c for c, cell in enumerate(ws[1], 1) if cell.value}

def pick(idx, *aliases):
    for a in aliases:
        if norm(a) in idx:
            return idx[norm(a)]
    return None

def ensure_col(ws, label):
    for c, cell in enumerate(ws[1], 1):
        if norm(cell.value) == norm(label):
            return c
    col = ws.max_column + 1
    ws.cell(1, col, label)
    return col


# ─── workbook loader ───────────────────────────────────────────────────────────
def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ld = wb["Localization Details"]
    mm = wb["Mention Mappings"]
    li = header_index(ld)
    mi = header_index(mm)

    c_id   = pick(li, "ID", "id")
    c_type = pick(li, "Type", "type")
    c_orig = pick(li, "Original Name", "original_name", "original name")
    c_loc  = pick(li, "Localized Name", "localised_name", "localized name", "localised name")
    c_fno  = pick(li, "First Name (Original)")
    c_lno  = pick(li, "Last Name (Original)")
    c_fnl  = pick(li, "First Name (Localized)", "First Name (Localised)")
    c_lnl  = pick(li, "Last Name (Localized)", "Last Name (Localised)")
    c_desc = pick(li, "Description", "description")

    chars: dict = {}
    for r in range(2, ld.max_row + 1):
        cid = ld.cell(r, c_id).value if c_id else None
        if not cid:
            continue
        chars[cid] = {
            "type": ld.cell(r, c_type).value if c_type else "",
            "orig": ld.cell(r, c_orig).value if c_orig else "",
            "loc":  ld.cell(r, c_loc).value if c_loc else "",
            "of":   norm(ld.cell(r, c_fno).value) if c_fno else "",
            "ol":   norm(ld.cell(r, c_lno).value) if c_lno else "",
            "lf":   (ld.cell(r, c_fnl).value or "").strip() if c_fnl else "",
            "ll":   (ld.cell(r, c_lnl).value or "").strip() if c_lnl else "",
            "desc": (ld.cell(r, c_desc).value or "") if c_desc else "",
        }

    m_id   = pick(mi, "ID", "id")
    m_orig = pick(mi, "Original Mention", "original_name", "original mention")
    m_loc  = pick(mi, "Localized Mention", "localised_name", "localized mention", "localised mention")

    mentions = []
    for r in range(2, mm.max_row + 1):
        orig = mm.cell(r, m_orig).value if m_orig else None
        if not orig:
            continue
        mentions.append({
            "row":  r,
            "cid":  mm.cell(r, m_id).value if m_id else None,
            "orig": orig,
            "loc":  mm.cell(r, m_loc).value if m_loc else "",
        })
    return wb, mm, chars, mentions


# ─── Layer 1: deterministic ────────────────────────────────────────────────────
def deterministic(cid, orig, loc, chars):
    """Return (status, suggestion, confidence, reason)."""
    c = chars.get(cid)
    if not loc:
        sug, ok = _name_parts(c, orig)
        return ("MISSING", sug if ok else "", 100 if ok else 0, "blank localized mention")
    if c and norm(orig) == norm(loc):
        sug, ok = _name_parts(c, orig)
        if ok and norm(sug) != norm(orig):
            return ("SOURCE_LEAK", sug, 100, "unchanged from source")
    sug, ok = _name_parts(c, orig)
    if ok:
        if norm(sug) == norm(loc):
            return ("VERIFIED", loc, 100, "matches character name parts")
        return ("MISMATCH", sug, 90, "differs from character name parts")
    return ("NEEDS_REVIEW", "", 0, "not resolvable deterministically")


def _de(name: str) -> str:
    return "d'" + name if name and name[0].lower() in "aeiou" else "de " + name


def _name_parts(c, orig):
    """Resolve using ONLY this character's own name parts + safe titles.
    Returns (suggestion, fully_resolved)."""
    if not c:
        return ("", False)
    out, unresolved, named = [], 0, False
    for tk in norm(orig).split():
        if c["of"] and tk == c["of"] and c["lf"]:
            out.append(c["lf"]); named = True
        elif c["ol"] and tk == c["ol"] and c["ll"]:
            out.append(c["ll"]); named = True
        elif tk.endswith("s") and len(tk) > 2:
            base = tk[:-1]
            if c["of"] and base == c["of"] and c["lf"]:
                out.append(_de(c["lf"])); named = True
            elif c["ol"] and base == c["ol"] and c["ll"]:
                out.append(_de(c["ll"])); named = True
            else:
                unresolved += 1
        elif tk in SAFE_TITLES:
            out.append(SAFE_TITLES[tk])
        elif tk in CONNECTORS:
            out.append(tk)
        else:
            unresolved += 1
    if unresolved or not named:
        return (" ".join(out), False)
    return (" ".join(out), True)


# ─── Layer 2: Madeye LLM ───────────────────────────────────────────────────────
MADEYE_MODEL = os.environ.get("MADEYE_MODEL", "claude-opus-4-8")
MADEYE_USER_EMAIL = os.environ.get("MADEYE_USER_EMAIL", "")


def _madeye_client(raise_on_missing: bool = False):
    from openai import OpenAI
    key  = os.environ.get("MADEYE_API_KEY", "")
    base = os.environ.get("MADEYE_BASE_URL", "")
    if not key:
        msg = "MADEYE_API_KEY not set"
        raise RuntimeError(msg) if raise_on_missing else sys.exit(msg)
    if not base:
        msg = "MADEYE_BASE_URL not set"
        raise RuntimeError(msg) if raise_on_missing else sys.exit(msg)
    return OpenAI(api_key=key, base_url=base, max_retries=2)


def llm_character(client, c, items, src, tgt, *, canon_ctx: str = "", user_email: str = ""):
    """One Madeye call resolving all unresolved mentions for a single character.

    canon_ctx — story canon text (authoritative over Description).
    """
    listing = "\n".join(
        f'  {i}. "{m["orig"]}"  (current: "{m["loc"]}")'
        for i, m in enumerate(items)
    )
    canon_block = ""
    if canon_ctx:
        canon_block = (
            "\nSTORY CANON (AUTHORITATIVE — overrides Description for family/surname/alias):\n"
            + canon_ctx[:3000] + "\n"
        )
    system = (
        f"You are a senior localization QA expert. Source {src} → target {tgt}.\n"
        + canon_block +
        "You verify how a character's name is rendered in each MENTION.\n"
        "Rules:\n"
        "- Canon overrides the Description for family, surname, and alias decisions.\n"
        "- Keep the SAME family/surname decisions made in the canonical localized name.\n"
        "- Resolve married names, aliases, diminutives, script errors via the Description.\n"
        "- Use natural French register (affectionate Tata/Tonton, honorifics Maitre etc.).\n"
        "- If the current localized mention contradicts the STORY CANON, set status CANON_CONFLICT.\n"
        "- confidence: 100 only if certain; lower if the description is silent.\n"
        'Respond ONLY as JSON: {"results":['
        '{"i":0,"suggestion":"...","confidence":95,"reason":"...","status":"LLM_FIX"}]}\n'
        'status must be one of: LLM_OK, LLM_FIX, CANON_CONFLICT'
    )
    user = (
        f'CHARACTER canonical: "{c["orig"]}" → "{c["loc"]}" '
        f'(first="{c["lf"]}", last="{c["ll"]}")\n\n'
        f'STORY DESCRIPTION:\n{c["desc"][:3000]}\n\n'
        f'MENTIONS to verify:\n{listing}\n'
    )
    email = user_email or MADEYE_USER_EMAIL
    resp = client.chat.completions.create(
        model=MADEYE_MODEL,
        messages=[{"role": "system", "content": system},
                  {"role": "user",   "content": user}],
        max_tokens=2000,
        temperature=0.1,
        extra_body={"metadata": {"user_email": email}},
    )
    raw = resp.choices[0].message.content or ""
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if not m:
        return {}
    try:
        data = json.loads(m.group(0))
    except json.JSONDecodeError:
        return {}
    return {r["i"]: r for r in data.get("results", []) if "i" in r}


# ─── write-back ────────────────────────────────────────────────────────────────
def write_back(mm, results):
    c_sug  = ensure_col(mm, "Suggested Localized Mention")
    c_conf = ensure_col(mm, "Confidence Score")
    c_stat = ensure_col(mm, "Status")
    c_rea  = ensure_col(mm, "Reason")
    for r in results:
        mm.cell(r["row"], c_sug,  r.get("suggestion", ""))
        mm.cell(r["row"], c_conf, f'{r["confidence"]}%' if r.get("confidence") else "")
        mm.cell(r["row"], c_stat, r["status"])
        mm.cell(r["row"], c_rea,  r.get("reason", ""))
        fill = FILLS.get(r["status"], "FFFFFF")
        for col in (c_sug, c_conf, c_stat):
            mm.cell(r["row"], col).fill = PatternFill("solid", fgColor=fill)
        if r.get("confidence") and r["confidence"] < 70:
            mm.cell(r["row"], c_conf).font = Font(bold=True)


# ─── apply human decisions (finalize) ─────────────────────────────────────────
def apply_decisions(mm, decisions: list[dict]) -> None:
    """Write accepted/edited values back into the Localized Mention column."""
    mi = header_index(mm)
    c_loc = pick(mi, "Localized Mention", "localised mention", "localized mention", "localised mention")
    if not c_loc:
        return
    for d in decisions:
        mm.cell(d["row"], c_loc, d["value"])


# ─── main orchestration (importable) ───────────────────────────────────────────
def run_verification(input_path: str, output_path: str, *,
                     use_llm: bool = False,
                     src: str = "de/en",
                     tgt: str = "fr",
                     corrections_lookup: dict | None = None,
                     canon_ctx: str = "",
                     user_email: str = "") -> dict:
    """Run the full verification pipeline and save the annotated workbook.

    Returns {"results": [...], "summary": {...}, "name_tokens": set}.
    results items: {row, cid, orig, loc, status, suggestion, confidence, reason}
    """
    wb, mm, chars, mentions = load(input_path)

    # Build name-token set for the proper-noun guard (used by corrections_store.add)
    name_tokens: set[str] = set()
    for c in chars.values():
        for tok in (norm(c.get("lf","")) + " " + norm(c.get("ll","")) + " " +
                    norm(c.get("orig",""))).split():
            if len(tok) > 2:
                name_tokens.add(tok)

    results: list[dict] = []

    for m in mentions:
        # Layer 0: learned corrections (highest priority)
        if corrections_lookup:
            key = (str(m["cid"]), norm(m["orig"]))
            if key in corrections_lookup:
                entry = corrections_lookup[key]
                results.append({
                    "row": m["row"], "cid": m["cid"],
                    "orig": m["orig"], "loc": m["loc"],
                    "status": entry["status_tag"],
                    "suggestion": entry["value"],
                    "confidence": 100,
                    "reason": "learned from human correction",
                })
                continue

        # Layer 1: deterministic
        st, sug, conf, rea = deterministic(m["cid"], m["orig"], m["loc"], chars)
        results.append({
            "row": m["row"], "cid": m["cid"],
            "orig": m["orig"], "loc": m["loc"],
            "status": st, "suggestion": sug, "confidence": conf, "reason": rea,
        })

    # Layer 2: LLM — only for unresolved rows
    if use_llm:
        client = _madeye_client(raise_on_missing=True)
        todo: dict[str, list] = defaultdict(list)
        for r in results:
            if r["status"] in ("NEEDS_REVIEW", "MISSING", "MISMATCH"):
                todo[r["cid"]].append(r)
        done = 0
        for cid, items in todo.items():
            c = chars.get(cid)
            if not c:
                continue
            try:
                res = llm_character(client, c, items, src, tgt,
                                    canon_ctx=canon_ctx, user_email=user_email)
            except Exception:
                continue
            for i, r in enumerate(items):
                rr = res.get(i)
                if not rr:
                    continue
                r["suggestion"] = rr.get("suggestion", r["suggestion"])
                r["confidence"] = int(rr.get("confidence", 0) or 0)
                r["reason"]     = rr.get("reason", "")[:200]
                llm_status = rr.get("status", "")
                if llm_status == "CANON_CONFLICT":
                    r["status"] = "CANON_CONFLICT"
                else:
                    r["status"] = "LLM_OK" if norm(r["suggestion"]) == norm(r["loc"]) else "LLM_FIX"
            done += 1

    write_back(mm, results)
    wb.save(output_path)

    counts: dict[str, int] = defaultdict(int)
    for r in results:
        counts[r["status"]] += 1

    auto = sum(
        1 for r in results
        if r["status"] in ("VERIFIED", "LLM_OK", "LEARNED") and r["confidence"] == 100
    )
    total = len(results)
    return {
        "results":     results,
        "name_tokens": name_tokens,
        "summary": {
            "total":        total,
            "auto_verified": auto,
            "needs_review": total - auto,
            "by_status":    dict(counts),
        },
    }


# ─── CLI wrapper ───────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Localization sheet verifier")
    ap.add_argument("input")
    ap.add_argument("-o", "--output", required=True)
    ap.add_argument("--llm", action="store_true", help="run LLM layer (needs MADEYE_API_KEY)")
    ap.add_argument("--src", default="de/en")
    ap.add_argument("--tgt", default="fr")
    args = ap.parse_args()

    corrections_lookup: dict = {}
    try:
        import corrections_store
        corrections_lookup = corrections_store.get_all()
        if corrections_lookup:
            print(f"Loaded {len(corrections_lookup)} learned corrections.")
    except Exception:
        pass

    if args.llm and not os.environ.get("MADEYE_USER_EMAIL"):
        sys.exit("MADEYE_USER_EMAIL not set — Madeye requires metadata.user_email.")

    result = run_verification(
        args.input, args.output,
        use_llm=args.llm,
        src=args.src,
        tgt=args.tgt,
        corrections_lookup=corrections_lookup,
        user_email=os.environ.get("MADEYE_USER_EMAIL", ""),
    )

    s = result["summary"]
    print(f"Loaded {s['total']} mentions.\n")
    print("Layer results:")
    for k, v in sorted(s["by_status"].items()):
        print(f"  {k:16s}: {v}")
    pct = f"{100*s['needs_review']/s['total']:.0f}" if s['total'] else "0"
    print(f"\nSaved → {args.output}")
    print(f"Human needs to review ~{s['needs_review']}/{s['total']} rows ({pct}%).")


if __name__ == "__main__":
    main()
