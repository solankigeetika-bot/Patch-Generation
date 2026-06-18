#!/usr/bin/env python3
"""
Evaluation harness: score the deterministic verifier against a human-reviewed
golden sheet.

The human-reviewed xlsx carries the verifier's ground truth in its
'Localization Issues' column. This harness:

  1. Reads the golden file.
  2. Strips the Issues column to reconstruct the "input" the tool would see.
  3. Runs every deterministic check from localisation_pipeline.
  4. Compares tool findings vs human findings at the (row, issue-type) level.
  5. Prints precision / recall / F1 per issue type and overall.

Usage:
    python eval_harness.py --golden TOLR_1200_Reviewed.xlsx \
        --source-lang de --target-lang fr
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict

import openpyxl

import localisation_pipeline as P


# Map the tool's internal `kind` tags to the human's UPPER_CASE issue tags.
TOOL_TO_HUMAN = {
    "missing_localised":             "MISSING_LOCALISATION",
    "same_first_name_collision":     "SAME_FIRST_NAME_COLLISION",
    "same_last_name_collision":      "SAME_LAST_NAME_COLLISION",
    "cultural_context_inappropriate":"CULTURAL_CONTEXT_INAPPROPRIATE",
    "entity_inconsistency":          "ENTITY_INCONSISTENCY",
    "cross_character_inconsistency": "CROSS_CHARACTER_INCONSISTENCY",
    "cross_batch_inconsistency":     "CROSS_CHARACTER_INCONSISTENCY",
    "mention_map_inconsistency":     "MENTION_MAP_INCONSISTENCY",
    "source_name_not_localised":     "SOURCE_NAME_NOT_LOCALIZED",
}

TAG_RE = re.compile(r"\b([A-Z][A-Z_]{4,})\s*:")


def human_tags(cell: str) -> set[str]:
    out = set()
    for seg in (cell or "").split(" | "):
        m = TAG_RE.match(seg.strip())
        if m:
            out.add(m.group(1))
    return out


def run(args) -> int:
    src, tgt = args.source_lang, args.target_lang
    wb = openpyxl.load_workbook(args.golden, data_only=True)

    ld_name = next((n for n in wb.sheetnames if "detail" in n.lower()), wb.sheetnames[0])
    mm_name = next((n for n in wb.sheetnames if "mention" in n.lower()), None)
    _, ld_rows = P.read_sheet(wb, ld_name)
    _, mm_rows = P.read_sheet(wb, mm_name) if mm_name else ([], [])

    print(f"Golden: {args.golden}", file=sys.stderr)
    print(f"  LD rows: {len(ld_rows)}   MM rows: {len(mm_rows)}\n", file=sys.stderr)

    # ── ground truth: row index -> set of human tags (LD only) ───────────────
    gold: dict[int, set[str]] = {}
    for i, row in enumerate(ld_rows):
        tags = human_tags(row.get("Localization Issues", ""))
        if tags:
            gold[i] = tags
    # strip the column so checks can't peek
    for row in ld_rows:
        row["Localization Issues"] = ""

    # ── run all deterministic checks ─────────────────────────────────────────
    universe = P.Universe()
    pred: dict[int, set[str]] = defaultdict(set)

    for i, row in enumerate(ld_rows):
        for f in P.review_proposition(row, universe, src, tgt):
            t = TOOL_TO_HUMAN.get(f.kind)
            if t:
                pred[i].add(t)

    for i, fs in P.check_name_collisions(ld_rows, src, tgt, mm_rows):
        for f in fs:
            t = TOOL_TO_HUMAN.get(f.kind)
            if t:
                pred[i].add(t)

    for i, fs in P.check_entity_consistency(ld_rows):
        for f in fs:
            t = TOOL_TO_HUMAN.get(f.kind)
            if t:
                pred[i].add(t)

    canonical, flat = P.check_flat_consistency(ld_rows)
    for i, fs in flat:
        for f in fs:
            t = TOOL_TO_HUMAN.get(f.kind)
            if t:
                pred[i].add(t)

    # ── Lever 1: MM checks attribute back to the character's LD row ──────────
    # The human records mention-map issues on the character's main LD row, so we
    # map each MM finding (keyed by Canonical Name) onto that LD row index.
    ld_by_orig: dict[str, int] = {}
    for i, row in enumerate(ld_rows):
        o = P.norm(row.get("Original Name", "")).lower()
        if o and o not in ld_by_orig:
            ld_by_orig[o] = i

    if mm_rows:
        for mm_i, fs in P.check_mention_consistency(mm_rows, ld_rows):
            cn = P.norm(mm_rows[mm_i].get("Canonical Name", "")).lower()
            ld_i = ld_by_orig.get(cn)
            if ld_i is None:
                continue
            for f in fs:
                t = TOOL_TO_HUMAN.get(f.kind)
                if t:
                    pred[ld_i].add(t)

    # ── score per type (row-level, LD) ───────────────────────────────────────
    types = sorted({t for ts in gold.values() for t in ts} |
                   {t for ts in pred.values() for t in ts})

    print(f"{'ISSUE TYPE':<34}{'TP':>5}{'FP':>5}{'FN':>5}"
          f"{'PREC':>7}{'REC':>7}{'F1':>7}")
    print("-" * 70)

    tot_tp = tot_fp = tot_fn = 0
    for t in types:
        gold_rows = {i for i, ts in gold.items() if t in ts}
        pred_rows = {i for i, ts in pred.items() if t in ts}
        tp = len(gold_rows & pred_rows)
        fp = len(pred_rows - gold_rows)
        fn = len(gold_rows - pred_rows)
        tot_tp += tp; tot_fp += fp; tot_fn += fn
        prec = tp / (tp + fp) if (tp + fp) else 0.0
        rec  = tp / (tp + fn) if (tp + fn) else 0.0
        f1   = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
        flag = "  ← tool can't do this yet" if (tp + fp) == 0 and fn else ""
        print(f"{t:<34}{tp:>5}{fp:>5}{fn:>5}{prec:>7.2f}{rec:>7.2f}{f1:>7.2f}{flag}")

    print("-" * 70)
    P_ = tot_tp / (tot_tp + tot_fp) if (tot_tp + tot_fp) else 0.0
    R_ = tot_tp / (tot_tp + tot_fn) if (tot_tp + tot_fn) else 0.0
    F_ = 2 * P_ * R_ / (P_ + R_) if (P_ + R_) else 0.0
    print(f"{'OVERALL (row-level)':<34}{tot_tp:>5}{tot_fp:>5}{tot_fn:>5}"
          f"{P_:>7.2f}{R_:>7.2f}{F_:>7.2f}")

    # recall over the human's TOTAL flagged rows (any type)
    gold_any = set(gold)
    pred_any = set(pred)
    row_recall = len(gold_any & pred_any) / len(gold_any) if gold_any else 0.0
    print(f"\nRows the human flagged:           {len(gold_any)}")
    print(f"Rows the tool also flagged (any): {len(gold_any & pred_any)} "
          f"({row_recall:.0%} of human rows touched)")
    print(f"Tool-only rows (review needed):   {len(pred_any - gold_any)}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Score verifier vs human golden sheet")
    ap.add_argument("--golden", required=True, help="human-reviewed xlsx")
    ap.add_argument("--source-lang", default="de")
    ap.add_argument("--target-lang", default="fr")
    return run(ap.parse_args())


if __name__ == "__main__":
    sys.exit(main())
